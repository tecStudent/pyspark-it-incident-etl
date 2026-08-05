import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path("data/raw/LW-DATASET.xlsx")
DEFAULT_OUTPUT = Path("data/raw/incidents.csv")
DEFAULT_SHEET = "Dataset Geral"


def extract_excel(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
) -> None:
    if not input_path.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {input_path}"
        )

    workbook = load_workbook(
        filename=input_path,
        read_only=True,
        data_only=True,
    )

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Aba '{sheet_name}' não encontrada. "
            f"Abas disponíveis: {workbook.sheetnames}"
        )

    worksheet = workbook[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)

    row_count = 0

    with output_path.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as csv_file:
        writer = csv.writer(csv_file)

        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(row)
            row_count += 1

    workbook.close()

    print(f"Arquivo de origem: {input_path}")
    print(f"Aba processada: {sheet_name}")
    print(f"Arquivo gerado: {output_path}")
    print(f"Registros extraídos: {row_count - 1}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrai dados do Excel para CSV."
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
    )

    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
    )

    args = parser.parse_args()

    extract_excel(
        input_path=args.input,
        output_path=args.output,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()